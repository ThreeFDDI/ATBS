#! /usr/local/bin/python3

import openpyxl, os
from openpyxl.utils import get_column_letter, column_index_from_string

os.chdir("/Users/JT/JTGIT/ATBS")

wb = openpyxl.load_workbook("ATBS-12_example.xlsx")

print(type(wb))

sheet = wb.get_sheet_by_name("Sheet3")

print(sheet)

print(type(sheet))

print(sheet.title)

anotherSheet = wb.active

print(anotherSheet)

print(anotherSheet.title)

#print(dir(anotherSheet))

a = anotherSheet["A1"]
b = anotherSheet["B1"]
c = anotherSheet["C1"]

print(a.value)
print(b.value)
print(c.value)

print()

print(b.row)
print(b.column)
print(b.coordinate)

print(get_column_letter(27))
print(get_column_letter(900))
print(column_index_from_string('AA'))
print(column_index_from_string('AHP'))

print(tuple(anotherSheet["A1":"C3"]))

print()
for row in anotherSheet["A1":"C3"]:
    for cell in row:
        print(cell.coordinate, cell.value)
    print("~~~ END OF ROW ~~~\n")